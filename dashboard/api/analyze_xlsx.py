"""Vercel Python function — analyse an insurer's Excel calculator and return RULES only.

POST /api/analyze_xlsx
  body: { "xlsx_url": "<signed Supabase Storage URL>", "insurer_name": "AIA" (optional) }
  ->  { insurer, source_file, structural_map, detection_log, rules, warnings }

Reuses the rate-compiler's Step-1 (structural map) + Step-2 (rule detection) logic,
trimmed to rules only — the rate NUMBERS come from the rate PDF, not here. Nothing is
guessed: undetected rules are returned as "NOT DETECTED — requires human input".

This is a vendored copy of tools/rate-compiler so it can be bundled as a standalone
serverless function; keep the detectors in sync if that tool changes.
"""
from __future__ import annotations

import datetime as _dt
import io
import json
import re
import urllib.request
from http.server import BaseHTTPRequestHandler

import openpyxl

NOT_DETECTED = "NOT DETECTED — requires human input"
NOTES_HINTS = ("note", "instruction", "readme", "read me", "guide", "import", "how to")
TEMPLATE_HINTS = ("input", "output", "quote", "calc", "template", "summary", "result", "form", "step")

_BAND_PATTERNS = [
    (re.compile(r"^\s*up\s*to\s*(\d{1,3})", re.I), lambda m: (0, int(m.group(1)))),
    (re.compile(r"^\s*(\d{1,3})\s*(?:-|to|–|—|~)\s*(\d{1,3})", re.I),
     lambda m: (int(m.group(1)), int(m.group(2)))),
    (re.compile(r"^\s*(\d{1,3})\s*(?:\+|and\s*(?:above|over|older)|&\s*above)", re.I),
     lambda m: (int(m.group(1)), 999)),
]


def parse_band(label):
    if not isinstance(label, str):
        return None
    for rx, fn in _BAND_PATTERNS:
        m = rx.search(label)
        if m:
            try:
                return fn(m)
            except (ValueError, IndexError):
                return None
    return None


def _jsonable(v):
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, (_dt.datetime, _dt.date, _dt.time)):
        return v.isoformat()
    return str(v)


# ── Dump the workbook (formulas + values + notes) from an in-memory xlsx ──────────
def dump_workbook(data: bytes) -> dict:
    wb_f = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    wb_v = openpyxl.load_workbook(io.BytesIO(data), data_only=True)

    sheets = [{"name": ws.title, "state": getattr(ws, "sheet_state", "visible"),
               "visible": getattr(ws, "sheet_state", "visible") == "visible"}
              for ws in wb_f.worksheets]

    formulas = {}
    for ws in wb_f.worksheets:
        m = {}
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if v is None:
                    continue
                if getattr(c, "data_type", None) == "f":
                    m[c.coordinate] = v if isinstance(v, str) else getattr(v, "text", str(v))
                elif isinstance(v, str) and v.startswith("="):
                    m[c.coordinate] = v
        if m:
            formulas[ws.title] = m

    values = {}
    for ws in wb_v.worksheets:
        m = {}
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None:
                    m[c.coordinate] = _jsonable(c.value)
        if m:
            values[ws.title] = m

    notes_chunks = []
    for s in sheets:
        if any(h in s["name"].lower() for h in NOTES_HINTS) and s["name"] in wb_v.sheetnames:
            ws = wb_v[s["name"]]
            notes_chunks.append(f"===== {s['name']} =====")
            for row in ws.iter_rows(values_only=True):
                line = " ".join(str(c) for c in row if c is not None).strip()
                if line:
                    notes_chunks.append(line)

    return {"all_sheets": sheets, "cell_formulas": formulas,
            "cell_values": values, "notes_text": "\n".join(notes_chunks)}


# ── STEP 1: structural map ────────────────────────────────────────────────────────
def classify_sheets(dump):
    out = []
    for s in dump["all_sheets"]:
        name, low = s["name"], s["name"].lower()
        cells = dump["cell_values"].get(name, {})
        bands = [r for r, v in cells.items() if parse_band(v)]
        numeric = sum(1 for v in cells.values() if isinstance(v, (int, float)) and not isinstance(v, bool))
        reasons = []
        if any(h in low for h in NOTES_HINTS):
            kind = "notes"; reasons.append(f"name matches notes hint ({low!r})")
        elif bands and numeric >= 4:
            kind = "rate_table"; reasons.append(f"{len(bands)} age-band label(s) beside {numeric} numeric cells")
        elif any(h in low for h in TEMPLATE_HINTS):
            kind = "template"; reasons.append(f"name matches input/output template hint ({low!r})")
        elif numeric == 0 and cells:
            kind = "notes"; reasons.append("text-only sheet")
        else:
            kind = "unknown"; reasons.append("no strong signal")
        if not s["visible"]:
            reasons.append(f"sheet is {s['state']}")
        out.append({"sheet": name, "state": s["state"], "classification": kind, "reasoning": "; ".join(reasons)})
    return out


# ── STEP 2: rule detection (rules only) ───────────────────────────────────────────
def _log(rule, value, source, confidence):
    return {"rule": rule, "value": value, "source_cell_or_text": source, "confidence": confidence}


def _all_text(dump):
    items = []
    for sheet, cells in dump["cell_values"].items():
        for ref, v in cells.items():
            if isinstance(v, str):
                items.append((sheet, ref, v))
    if dump.get("notes_text"):
        items.append(("<notes_text>", "-", dump["notes_text"]))
    return items


def _search(dump, patterns):
    for sheet, ref, v in _all_text(dump):
        for rx in patterns:
            m = rx.search(v)
            if m:
                snip = v if len(v) <= 120 else v[max(0, m.start() - 30): m.start() + 90]
                where = f"{sheet}!{ref}" if ref != "-" else sheet
                return snip.strip(), where
    return None


def detect_age_basis(dump):
    nb = _search(dump, [re.compile(r"age\s+next\s+birthday|next\s+birthday|\bANB\b", re.I)])
    lb = _search(dump, [re.compile(r"age\s+last\s+birthday|last\s+birthday|\bALB\b", re.I)])
    if nb and not lb:
        return _log("age_basis", "next birthday", f"{nb[1]}: {nb[0]!r}", "high")
    if lb and not nb:
        return _log("age_basis", "last birthday", f"{lb[1]}: {lb[0]!r}", "high")
    if nb and lb:
        return _log("age_basis", NOT_DETECTED, f"both found — {nb[1]}:{nb[0]!r} AND {lb[1]}:{lb[0]!r}", "low")
    return _log("age_basis", NOT_DETECTED, "no 'last/next birthday' text found", "n/a")


def detect_gst(dump):
    for sheet, cells in dump["cell_formulas"].items():
        for ref, f in cells.items():
            m = re.search(r"/\s*(1\.0[789])\b", f)
            if m:
                return _log("gst_treatment", {"treatment": "inclusive", "conversion_factor": float(m.group(1))},
                            f"{sheet}!{ref}: {f!r}", "high")
    inc = _search(dump, [re.compile(r"inclusive of\s+gst|gst[- ]inclusive|incl\.?\s*gst|net of gst", re.I)])
    if inc:
        return _log("gst_treatment", {"treatment": "inclusive", "conversion_factor": None}, f"{inc[1]}: {inc[0]!r}", "medium")
    exc = _search(dump, [re.compile(r"exclusive of\s+gst|gst[- ]exclusive|excl\.?\s*gst|before gst|excluding gst", re.I)])
    if exc:
        return _log("gst_treatment", {"treatment": "exclusive", "conversion_factor": None}, f"{exc[1]}: {exc[0]!r}", "medium")
    return _log("gst_treatment", NOT_DETECTED, "no GST factor or inclusive/exclusive text found", "n/a")


def detect_group_size_discount(dump):
    hit = _search(dump, [
        re.compile(r"(group|head\s*count|lives|members?|employees?|\bpax\b)\D{0,25}(discount|loading|rebate)", re.I),
        re.compile(r"(discount|loading|rebate)\D{0,25}(group|head\s*count|lives|members?|employees?)", re.I),
    ])
    if hit:
        return _log("group_size_discount", "detected — confirm the tier factors (not auto-parsed)", f"{hit[1]}: {hit[0]!r}", "low")
    return _log("group_size_discount", NOT_DETECTED, "no headcount discount/loading text found", "n/a")


def detect_rider_dependencies(dump):
    hits = []
    rx = re.compile(r"([A-Z][A-Za-z0-9+/&\- ]{1,40}?)\s+(?:must be taken with|requires|only with|rider to|is a\s+[A-Za-z ]*rider|attach(?:ed)? to|bundle[sd]? (?:in|with))\s+([A-Z][A-Za-z0-9+/&\- ]{1,40})", re.I)
    for sheet, ref, v in _all_text(dump):
        m = rx.search(v)
        if m:
            hits.append({"coverage": m.group(1).strip(), "requires": m.group(2).strip(), "source": f"{sheet}!{ref}"})
    if hits:
        return _log("rider_dependencies", hits, "; ".join(h["source"] for h in hits), "medium")
    return _log("rider_dependencies", NOT_DETECTED, "no rider dependency phrasing found", "n/a")


def detect_renewal_only_bands(dump):
    hits = []
    for sheet, ref, v in _all_text(dump):
        if re.search(r"renewal\s*only|renewal\s*basis|not.*new business|renewals?\s+only", v, re.I):
            hits.append({"text": v.strip()[:100], "band": parse_band(v), "source": f"{sheet}!{ref}"})
    if hits:
        return _log("renewal_only_bands", hits, "; ".join(h["source"] for h in hits), "medium")
    return _log("renewal_only_bands", NOT_DETECTED, "no 'renewal only' text found", "n/a")


def detect_occupation_class(dump):
    classes, src = set(), None
    for sheet, ref, v in _all_text(dump):
        for m in re.finditer(r"\b(?:class|occupation(?:al)?\s*class)\s*([1-4])\b", v, re.I):
            classes.add(int(m.group(1)))
            src = src or f"{sheet}!{ref}: {v.strip()[:60]!r}"
    if classes:
        rule = _search(dump, [re.compile(r"class\s*[1-4].{0,60}(eligib|exclud|not (?:eligible|covered)|load)", re.I)])
        note = f"; rule at {rule[1]}: {rule[0]!r}" if rule else "; no explicit eligibility/loading rule — effect NOT confirmed"
        return _log("occupation_class_rules",
                    {"classes_present": sorted(classes), "rule": rule[0] if rule else NOT_DETECTED},
                    (src or "") + note, "medium" if rule else "low")
    return _log("occupation_class_rules", NOT_DETECTED, "no occupation Class 1-4 references found", "n/a")


DETECTORS = [detect_age_basis, detect_gst, detect_group_size_discount,
             detect_rider_dependencies, detect_renewal_only_bands, detect_occupation_class]


def analyze(data: bytes, insurer_name=None) -> dict:
    dump = dump_workbook(data)
    detection_log = [fn(dump) for fn in DETECTORS]
    rules = {e["rule"]: e["value"] for e in detection_log}
    warnings = []
    if not any(s["classification"] == "rate_table" for s in classify_sheets(dump)):
        warnings.append("No obvious rate-table sheet found — this looks like a pure calculator/template (rules-only is fine).")
    return {
        "insurer": insurer_name or None,
        "structural_map": classify_sheets(dump),
        "detection_log": detection_log,
        "rules": rules,
        "warnings": warnings,
    }


# ── Vercel handler ────────────────────────────────────────────────────────────────
class handler(BaseHTTPRequestHandler):
    def _send(self, code, obj):
        body = json.dumps(obj).encode("utf-8")
        self.send_response(code)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        # Health check — a browser visit confirms the Python function is live.
        self._send(200, {"ok": True, "service": "analyze_xlsx", "usage": "POST { xlsx_url, insurer_name? }"})

    def do_POST(self):
        try:
            length = int(self.headers.get("Content-Length", 0))
            payload = json.loads(self.rfile.read(length) or b"{}")
            url = payload.get("xlsx_url")
            if not url:
                return self._send(400, {"error": "xlsx_url required"})
            with urllib.request.urlopen(url, timeout=30) as resp:  # noqa: S310 (trusted signed URL)
                data = resp.read()
            result = analyze(data, payload.get("insurer_name"))
            self._send(200, result)
        except Exception as e:  # noqa: BLE001 — surface the message to the caller
            self._send(500, {"error": f"{type(e).__name__}: {e}"})
