"""Vercel Python function — DUMP an insurer Excel calculator's structure.

POST /api/pm_dump
  body: { "xlsx_url": "<signed Supabase Storage URL>" }
  ->  { sheets, validations, previews, notes_text }

Mechanical only (openpyxl) — no guessing. Produces the material an LLM needs to PROPOSE a
cell-map profile (which cells are inputs / plan-selections / outputs / totals) and that the
review UI shows a human to confirm it:
  - sheets:      name/state/visible/dims
  - validations: dropdown domains per cell range (Employee/Dependent, Plan 1..5, 0%/20%, ...)
  - previews:    a top-rows text grid (headers + first data rows), formula samples, SUM cells
  - notes_text:  concatenated text from any notes/instruction sheets (business rules)
"""
from __future__ import annotations

import datetime as _dt
import io
import json
import re
import urllib.request
from http.server import BaseHTTPRequestHandler

import openpyxl

NOTES_HINTS = ("note", "instruction", "readme", "read me", "guide", "how to")
# Sheets whose name suggests the actual rate/premium/plan matrix — filled into `values` first, so a
# large scratch/working sheet earlier in the file can't burn the global cap before the real rate
# table (possibly hidden, possibly later in the workbook) is reached.
RATE_HINTS = ("rate", "premium", "price", "pricing", "plan", "benefit", "schedule", "cover")
SCRATCH_HINTS = ("working", "calc", "scratch", "temp", "draft", "helper", "summary")
MAX_TOP_ROWS = 20
MAX_FORMULA_SAMPLES = 24
MAX_SUM_CELLS = 40
CELL_TEXT_CAP = 60
MAX_VALUES_PER_SHEET = 3000
MAX_VALUES_TOTAL = 12000


def _sheet_priority(name: str) -> int:
    """Lower = filled first. Rate/premium/plan-like names win, scratch/working sheets lose,
    everything else is neutral — order among equal-priority sheets stays file order (stable sort)."""
    n = name.lower()
    if any(h in n for h in RATE_HINTS):
        return 0
    if any(h in n for h in SCRATCH_HINTS):
        return 2
    return 1


def _short(v):
    if v is None:
        return None
    if isinstance(v, (_dt.datetime, _dt.date, _dt.time)):
        return v.isoformat()
    s = str(v)
    return s if len(s) <= CELL_TEXT_CAP else s[:CELL_TEXT_CAP] + "…"


def _formula_text(cell):
    v = cell.value
    if getattr(cell, "data_type", None) == "f":
        return v if isinstance(v, str) else getattr(v, "text", str(v))
    if isinstance(v, str) and v.startswith("="):
        return v
    return None


def _parse_list_options(formula1: str):
    """A list data-validation stores its options either inline ('"a, b, c"') or as a range ref.
    Return the inline options; range-backed lists return None (options live in cells)."""
    if not formula1:
        return None
    s = formula1.strip()
    if s.startswith('"') and s.endswith('"'):
        s = s[1:-1]
        return [o.strip() for o in s.split(",") if o.strip()]
    return None  # range reference — caller can resolve if needed


def dump(data: bytes) -> dict:
    wb_f = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    wb_v = openpyxl.load_workbook(io.BytesIO(data), data_only=True)

    sheets = [{"name": ws.title, "state": getattr(ws, "sheet_state", "visible"),
               "visible": getattr(ws, "sheet_state", "visible") == "visible",
               "max_row": ws.max_row, "max_col": ws.max_column}
              for ws in wb_f.worksheets]

    validations: dict = {}
    for ws in wb_f.worksheets:
        dvs = []
        try:
            for dv in ws.data_validations.dataValidation:
                dvs.append({
                    "range": str(dv.sqref),
                    "type": dv.type,
                    "options": _parse_list_options(dv.formula1) if dv.type == "list" else None,
                    "formula1": dv.formula1 if dv.type == "list" else None,
                })
        except Exception:
            pass
        if dvs:
            validations[ws.title] = dvs

    previews: dict = {}
    for ws in wb_f.worksheets:
        wsv = wb_v[ws.title]
        # top rows: headers + first data rows (formula cells shown as "=<formula>")
        top_rows = []
        for r in range(1, min(MAX_TOP_ROWS, ws.max_row) + 1):
            cells = {}
            for c in ws[r]:
                fx = _formula_text(c)
                if fx is not None:
                    cells[c.coordinate[:-len(str(r))]] = fx[:CELL_TEXT_CAP]
                else:
                    val = wsv[c.coordinate].value
                    if val is not None:
                        cells[c.coordinate[:-len(str(r))]] = _short(val)
            if cells:
                top_rows.append({"row": r, "cells": cells})

        formula_samples, sum_cells = [], []
        for row in ws.iter_rows():
            for c in row:
                fx = _formula_text(c)
                if fx is None:
                    continue
                if len(formula_samples) < MAX_FORMULA_SAMPLES:
                    formula_samples.append({"cell": c.coordinate, "formula": fx})
                if "SUM(" in fx.upper() and len(sum_cells) < MAX_SUM_CELLS:
                    sum_cells.append({"cell": c.coordinate, "formula": fx})

        # Grand-total hints: a "Total" text label + the cell to its right (which usually holds
        # the grand-total formula, e.g. Y118 "Total :" -> Z118 =N116+Q116+...).
        total_hints = []
        for row in ws.iter_rows():
            for c in row:
                v = wsv[c.coordinate].value
                if isinstance(v, str) and "total" in v.lower() and len(v) <= 40:
                    right = ws.cell(c.row, c.column + 1)
                    rfx = _formula_text(right)
                    rval = wsv[right.coordinate].value
                    if rfx or rval is not None:
                        total_hints.append({"label_cell": c.coordinate, "label": _short(v),
                                            "value_cell": right.coordinate, "formula": rfx})
            if len(total_hints) >= 12:
                break

        previews[ws.title] = {"top_rows": top_rows, "formula_samples": formula_samples,
                              "sum_cells": sum_cells, "total_hints": total_hints}

    notes_chunks = []
    for s in sheets:
        if any(h in s["name"].lower() for h in NOTES_HINTS):
            ws = wb_v[s["name"]]
            notes_chunks.append(f"===== {s['name']} =====")
            for row in ws.iter_rows(values_only=True):
                line = " ".join(str(c) for c in row if c is not None).strip()
                if line:
                    notes_chunks.append(line)

    # Full cached values per sheet (for the pricing extractor to read every rate, and for the UI to
    # let a human open any sheet). Capped so the payload stays reasonable — filled in relevance
    # order (see _sheet_priority) so a rate-like sheet is never the one left truncated because a
    # scratch/working sheet happened to come first in the file.
    values = {}
    total = 0
    ordered = sorted(wb_v.worksheets, key=lambda ws: _sheet_priority(ws.title))
    for ws in ordered:
        m = {}
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None:
                    m[c.coordinate] = _short(c.value)
                    total += 1
            if len(m) >= MAX_VALUES_PER_SHEET or total >= MAX_VALUES_TOTAL:
                m["_truncated"] = "true"
                break
        if m:
            values[ws.title] = m
        if total >= MAX_VALUES_TOTAL:
            break

    return {"sheets": sheets, "validations": validations, "previews": previews,
            "values": values, "notes_text": "\n".join(notes_chunks)[:8000]}


# ── Vercel handler ────────────────────────────────────────────────────────────────
class handler(BaseHTTPRequestHandler):
    def _send(self, code, obj):
        body = json.dumps(obj).encode("utf-8")
        self.send_response(code)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        self._send(200, {"ok": True, "service": "pm_dump", "usage": "POST { xlsx_url }"})

    def do_POST(self):
        try:
            length = int(self.headers.get("Content-Length", 0))
            payload = json.loads(self.rfile.read(length) or b"{}")
            url = payload.get("xlsx_url")
            if not url:
                return self._send(400, {"error": "xlsx_url required"})
            with urllib.request.urlopen(url, timeout=30) as resp:  # noqa: S310 (trusted signed URL)
                data = resp.read()
            self._send(200, dump(data))
        except Exception as e:  # noqa: BLE001
            self._send(500, {"error": f"{type(e).__name__}: {e}"})
