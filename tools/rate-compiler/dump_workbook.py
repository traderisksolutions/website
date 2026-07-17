#!/usr/bin/env python3
"""dump_workbook.py — openpyxl -> structured JSON dump for the rate compiler.

Reads an insurance premium-calculator .xlsx and emits the exact JSON shape that
compile_rates.py consumes:

    {
      "source_file":   str,
      "all_sheets":    [ {name, state, visible} ],   # incl. hidden / veryHidden
      "cell_formulas": {sheet: {cell_ref: formula_string}},   # non-empty formula cells
      "cell_values":   {sheet: {cell_ref: computed_value}},   # non-empty value cells
      "named_ranges":  {name: {ref, hidden}},
      "has_vba":       bool,
      "vba_modules":   [module_name],
      "notes_text":    str    # concatenated text from Notes / instructions sheets
    }

openpyxl can't recompute formulas, so we read the workbook TWICE:
  - data_only=False -> formulas (the "=..." strings)
  - data_only=True  -> the values Excel last cached on save

Usage:
    python3 dump_workbook.py path/to/calculator.xlsx [-o dump.json]
    python3 dump_workbook.py path/to/calculator.xlsx           # -> stdout
"""
from __future__ import annotations

import argparse
import datetime as _dt
import json
import os
import sys
import zipfile

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - environment guard
    sys.stderr.write(
        "openpyxl is required: pip install -r requirements.txt "
        "(or: python3 -m pip install openpyxl)\n"
    )
    raise SystemExit(2)


# Sheet-name fragments that mark a sheet as free-text notes / instructions rather
# than rate data. Matched case-insensitively as substrings.
NOTES_HINTS = ("note", "instruction", "readme", "read me", "guide", "import", "how to")


def _jsonable(v):
    """Coerce an openpyxl cell value into something json.dumps can serialise."""
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, (_dt.datetime, _dt.date, _dt.time)):
        return v.isoformat()
    if isinstance(v, _dt.timedelta):
        return v.total_seconds()
    return str(v)


def _sheet_states(wb) -> list[dict]:
    """Every sheet with its visibility state. openpyxl exposes 'visible',
    'hidden', and 'veryHidden' via ws.sheet_state."""
    out = []
    for ws in wb.worksheets:
        state = getattr(ws, "sheet_state", "visible")
        out.append({"name": ws.title, "state": state, "visible": state == "visible"})
    return out


def _collect_formulas(wb_formulas) -> dict:
    """{sheet: {cell_ref: formula_string}} for every cell whose value is a formula.

    In data_only=False mode a formula cell's .value is the formula string starting
    with '=' (openpyxl also flags data_type == 'f'). Array formulas surface as
    ArrayFormula objects — we keep their .text."""
    formulas: dict[str, dict] = {}
    for ws in wb_formulas.worksheets:
        sheet_map: dict[str, str] = {}
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if val is None:
                    continue
                text = None
                if getattr(cell, "data_type", None) == "f":
                    text = val if isinstance(val, str) else getattr(val, "text", str(val))
                elif isinstance(val, str) and val.startswith("="):
                    text = val
                else:
                    # ArrayFormula / DataTableFormula objects have a .text / .ref
                    t = getattr(val, "text", None)
                    if t is not None:
                        text = t
                if text is not None:
                    sheet_map[cell.coordinate] = text
        if sheet_map:
            formulas[ws.title] = sheet_map
    return formulas


def _collect_values(wb_values) -> dict:
    """{sheet: {cell_ref: computed_value}} for every non-empty cell (data_only)."""
    values: dict[str, dict] = {}
    for ws in wb_values.worksheets:
        sheet_map: dict[str, object] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                sheet_map[cell.coordinate] = _jsonable(cell.value)
        if sheet_map:
            values[ws.title] = sheet_map
    return values


def _collect_named_ranges(wb) -> dict:
    """{name: {ref, hidden}}. Workbook-global defined names; openpyxl exposes
    `hidden` on each DefinedName."""
    named: dict[str, dict] = {}
    try:
        items = wb.defined_names.items()  # openpyxl >= 3.1: dict-like
    except AttributeError:  # pragma: no cover - older openpyxl
        items = [(dn.name, dn) for dn in wb.defined_names.definedName]
    for name, dn in items:
        ref = getattr(dn, "attr_text", None) or getattr(dn, "value", None)
        named[name] = {"ref": ref, "hidden": bool(getattr(dn, "hidden", False))}
    return named


def _detect_vba(path: str) -> tuple[bool, list[str]]:
    """An .xlsx/.xlsm is a zip. Macros live in xl/vbaProject.bin — its presence is
    the reliable signal. We can't cleanly parse module names without oletools, so we
    report the container and any obvious module streams we can see."""
    try:
        with zipfile.ZipFile(path) as zf:
            names = zf.namelist()
    except zipfile.BadZipFile:
        return False, []
    has_vba = any(n.lower() == "xl/vbaproject.bin" for n in names)
    modules: list[str] = []
    if has_vba:
        # Best-effort: surface the binary blob so downstream knows macros exist.
        # Full module extraction needs oletools; we deliberately don't guess names.
        modules = ["vbaProject.bin (module names require oletools to enumerate)"]
    return has_vba, modules


def _collect_notes(wb_values, sheet_states: list[dict]) -> str:
    """Concatenate text from sheets whose name looks like notes/instructions."""
    chunks: list[str] = []
    note_sheets = [
        s["name"] for s in sheet_states
        if any(h in s["name"].lower() for h in NOTES_HINTS)
    ]
    for name in note_sheets:
        if name not in wb_values.sheetnames:
            continue
        ws = wb_values[name]
        chunks.append(f"===== {name} =====")
        for row in ws.iter_rows(values_only=True):
            line = " ".join(str(c) for c in row if c is not None).strip()
            if line:
                chunks.append(line)
    return "\n".join(chunks)


def dump(path: str) -> dict:
    # Two loads: formulas (data_only=False) and cached values (data_only=True).
    wb_f = openpyxl.load_workbook(path, data_only=False, keep_vba=path.lower().endswith(".xlsm"))
    wb_v = openpyxl.load_workbook(path, data_only=True)

    sheet_states = _sheet_states(wb_f)
    has_vba, vba_modules = _detect_vba(path)

    result = {
        "source_file": os.path.basename(path),
        "all_sheets": sheet_states,
        "cell_formulas": _collect_formulas(wb_f),
        "cell_values": _collect_values(wb_v),
        "named_ranges": _collect_named_ranges(wb_f),
        "has_vba": has_vba,
        "vba_modules": vba_modules,
        "notes_text": _collect_notes(wb_v, sheet_states),
    }
    return result


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="Dump an .xlsx to the rate-compiler JSON shape.")
    ap.add_argument("xlsx", help="Path to the .xlsx / .xlsm calculator")
    ap.add_argument("-o", "--out", help="Write JSON here (default: stdout)")
    args = ap.parse_args(argv)

    if not os.path.isfile(args.xlsx):
        sys.stderr.write(f"No such file: {args.xlsx}\n")
        return 2

    data = dump(args.xlsx)
    text = json.dumps(data, indent=2, ensure_ascii=False)
    if args.out:
        with open(args.out, "w", encoding="utf-8") as fh:
            fh.write(text)
        sheets = len(data["all_sheets"])
        hidden = sum(1 for s in data["all_sheets"] if not s["visible"])
        sys.stderr.write(
            f"Wrote {args.out}: {sheets} sheets ({hidden} hidden), "
            f"vba={data['has_vba']}\n"
        )
    else:
        print(text)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
